"""
Generate Teachific Quiz Import Template ZIP
Creates: QuizImportTemplate.zip containing:
  - QuizImportTemplate.xlsx (Instructions, Questions, Template sheets)
  - media/ folder with sample placeholder images
"""

import os
import zipfile
import io
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ─── Colors ───────────────────────────────────────────────────────────────────
TEAL       = "FF189AA1"
TEAL_LIGHT = "FFE0F7F8"
AQUA       = "FF4AD9E0"
DARK       = "FF1A2332"
WHITE      = "FFFFFFFF"
GRAY_BG    = "FFF5F5F5"
GRAY_BORDER= "FFD0D0D0"
YELLOW     = "FFFFF3CD"
GREEN_LIGHT= "FFD4EDDA"
RED_LIGHT  = "FFF8D7DA"

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_border(color=GRAY_BORDER):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(size=11, bold=True, color=WHITE):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color="FF333333"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

# ─── Sheet 1: Instructions ────────────────────────────────────────────────────
def build_instructions_sheet(wb):
    ws = wb.create_sheet("Instructions", 0)
    ws.sheet_view.showGridLines = False

    # Title row
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "Teachific™ Quiz Import Template — Instructions"
    c.fill = make_fill(TEAL)
    c.font = Font(name="Calibri", size=16, bold=True, color=WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # Subtitle
    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = "Import quizzes into Teachific by uploading a ZIP file containing your Excel spreadsheet and optional media files."
    c.fill = make_fill(TEAL_LIGHT)
    c.font = Font(name="Calibri", size=11, color=DARK)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    row = 4

    def section(title, color=TEAL):
        nonlocal row
        ws.merge_cells(f"A{row}:H{row}")
        c = ws[f"A{row}"]
        c.value = title
        c.fill = make_fill(color)
        c.font = Font(name="Calibri", size=12, bold=True, color=WHITE)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 24
        row += 1

    def para(text, indent=1, fill=None, bold=False):
        nonlocal row
        ws.merge_cells(f"A{row}:H{row}")
        c = ws[f"A{row}"]
        c.value = text
        c.font = Font(name="Calibri", size=10, bold=bold, color=DARK)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=indent)
        if fill:
            c.fill = make_fill(fill)
        ws.row_dimensions[row].height = 18
        row += 1

    def table_header(cols):
        nonlocal row
        for i, (col_letter, text) in enumerate(cols):
            c = ws[f"{col_letter}{row}"]
            c.value = text
            c.fill = make_fill(DARK)
            c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = make_border(DARK)
        ws.row_dimensions[row].height = 20
        row += 1

    def table_row(cols, alt=False):
        nonlocal row
        fill = GRAY_BG if alt else WHITE
        for col_letter, text in cols:
            c = ws[f"{col_letter}{row}"]
            c.value = text
            c.font = body_font()
            c.fill = make_fill(fill)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
            c.border = make_border()
        ws.row_dimensions[row].height = 18
        row += 1

    def blank():
        nonlocal row
        row += 1

    # ── Section 1: ZIP Structure ──────────────────────────────────────────────
    section("1. ZIP File Structure")
    para("Your import file must be a .zip archive with the following structure:")
    blank()
    para("  MyQuiz.zip", indent=2, fill=GRAY_BG, bold=True)
    para("  ├── QuizImportTemplate.xlsx   ← Your questions spreadsheet (required)", indent=3)
    para("  └── media/                    ← Folder containing your media files (optional)", indent=3)
    para("       ├── question_image.jpg", indent=4)
    para("       ├── intro_video.mp4", indent=4)
    para("       └── audio_clip.mp3", indent=4)
    blank()
    para("• If you have no media, you can upload the .xlsx file directly (without a ZIP).", fill=YELLOW)
    para("• Media files must be placed inside the media/ folder at the root of the ZIP.", fill=YELLOW)
    para("• Reference media files in the Image, Video, or Audio columns using the path: media/filename.ext", fill=YELLOW)
    blank()

    # ── Section 2: Spreadsheet Sheets ────────────────────────────────────────
    section("2. Spreadsheet Sheets")
    para("The Excel file must contain a sheet named 'Questions' with your quiz data.")
    para("Optionally include a 'Template' sheet (provided in this file) as a reference — it will be ignored during import.")
    blank()

    # ── Section 3: Column Reference ──────────────────────────────────────────
    section("3. Column Reference")
    blank()
    table_header([
        ("A", "Column"), ("B", "Header"), ("C", "Required?"), ("D", "Description")
    ])
    cols = [("A",""), ("B",""), ("C",""), ("D","")]
    ref = [
        ("A", "Question Type", "Required", "Type code: TF, MC, MR, TI, MG, SEQ, NUMG, IS (see Section 4)"),
        ("B", "Question Text", "Required", "The full question text. Supports multi-line (use Alt+Enter in Excel)."),
        ("C", "Image", "Optional", "Relative path to an image file in the media/ folder. Example: media/diagram.jpg"),
        ("D", "Video", "Optional", "Relative path to a video file in the media/ folder. Example: media/intro.mp4"),
        ("E", "Audio", "Optional", "Relative path to an audio file in the media/ folder. Example: media/narration.mp3"),
        ("F–O", "Answer 1–10", "Varies", "Answer choices. Prefix correct answers with * (e.g. *Paris). See Section 4 for type-specific rules."),
        ("P", "Correct Feedback", "Optional", "Message shown when the learner answers correctly."),
        ("Q", "Incorrect Feedback", "Optional", "Message shown when the learner answers incorrectly."),
        ("R", "Points", "Optional", "Point value for this question (numeric). Leave blank for default (1 point)."),
    ]
    for i, (col, header, req, desc) in enumerate(ref):
        ws.merge_cells(f"B{row}:B{row}")
        ws.merge_cells(f"D{row}:H{row}")
        c_a = ws[f"A{row}"]
        c_b = ws[f"B{row}"]
        c_c = ws[f"C{row}"]
        c_d = ws[f"D{row}"]
        fill = GRAY_BG if i % 2 == 0 else WHITE
        for c, v in [(c_a, col), (c_b, header), (c_c, req), (c_d, desc)]:
            c.value = v
            c.font = body_font()
            c.fill = make_fill(fill)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
            c.border = make_border()
        ws.row_dimensions[row].height = 20
        row += 1
    blank()

    # ── Section 4: Question Types ─────────────────────────────────────────────
    section("4. Question Types")
    blank()
    table_header([("A", "Code"), ("B", "Type"), ("C", "Answer Format"), ("D", "Example")])
    types = [
        ("TF",   "True / False",         "Answer 1 = *True or *False (prefix * on correct)",                          "*True  |  False"),
        ("MC",   "Multiple Choice",       "One correct answer, prefix * on the correct option",                         "*Paris  |  London  |  Berlin"),
        ("MR",   "Multiple Response",     "Multiple correct answers, prefix * on each correct option",                   "*Red  |  *Blue  |  Green  |  Yellow"),
        ("TI",   "Short Answer",          "List all accepted answers in Answer 1–10 (no * needed)",                      "Gone With the Wind  |  gone with the wind"),
        ("MG",   "Matching",              "Each answer is 'Premise|Response' (pipe-separated)",                          "France|Paris  |  Germany|Berlin"),
        ("SEQ",  "Sequence / Order",      "List items in the correct order in Answer 1–N",                               "First  |  Second  |  Third  |  Fourth"),
        ("NUMG", "Numeric",               "Use =X for exact, X..Y for range, >X, <X, >=X, <=X, !=X",                    "=42  or  40..44  or  >10"),
        ("IS",   "Info Slide",            "Answer 1 = descriptive text (no correct/incorrect)",                          "This section covers Chapter 3."),
    ]
    for i, (code, qtype, fmt, ex) in enumerate(types):
        fill = GRAY_BG if i % 2 == 0 else WHITE
        ws.merge_cells(f"C{row}:C{row}")
        ws.merge_cells(f"D{row}:F{row}")
        ws.merge_cells(f"G{row}:H{row}")
        for col_letter, val in [("A", code), ("B", qtype), ("C", fmt), ("G", ex)]:
            c = ws[f"{col_letter}{row}"]
            c.value = val
            c.font = body_font(bold=(col_letter == "A"))
            c.fill = make_fill(fill)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
            c.border = make_border()
        ws.row_dimensions[row].height = 22
        row += 1
    blank()

    # ── Section 5: Supported Media Formats ───────────────────────────────────
    section("5. Supported Media Formats")
    blank()
    para("Images:  .jpg, .jpeg, .png, .gif, .webp, .svg")
    para("Video:   .mp4, .webm, .mov, .avi")
    para("Audio:   .mp3, .wav, .ogg, .m4a, .aac")
    blank()
    para("Maximum file size per media item: 50 MB", fill=YELLOW)
    para("Maximum total ZIP size: 200 MB", fill=YELLOW)
    blank()

    # ── Section 6: Tips ───────────────────────────────────────────────────────
    section("6. Tips & Common Mistakes")
    blank()
    para("✓  The 'Questions' sheet name is case-sensitive — it must be exactly 'Questions'.")
    para("✓  Do not include a header row in the Questions sheet — data starts on row 1.")
    para("✓  Media paths are case-sensitive on most systems. Use lowercase filenames.")
    para("✓  For Matching questions, the pipe character | separates premise from response.")
    para("✓  For True/False, the correct answer must be exactly 'True' or 'False' (with * prefix).")
    para("✓  Leave the Points column blank to use the default value of 1 point.")
    para("✓  Info Slides (IS) are not graded — they display content only.")
    blank()
    para("Need help? Visit teachific.app/support or contact your administrator.", fill=TEAL_LIGHT)

    # Column widths
    set_col_width(ws, "A", 12)
    set_col_width(ws, "B", 28)
    set_col_width(ws, "C", 18)
    set_col_width(ws, "D", 50)
    set_col_width(ws, "E", 20)
    set_col_width(ws, "F", 20)
    set_col_width(ws, "G", 30)
    set_col_width(ws, "H", 20)

    return ws


# ─── Sheet 2: Questions (sample data) ────────────────────────────────────────
def build_questions_sheet(wb):
    ws = wb.create_sheet("Questions")
    ws.sheet_view.showGridLines = False

    # No header row — data starts at row 1 per the import spec
    rows = [
        # TF
        ["TF",
         "The Great Wall of China is visible from space with the naked eye.",
         "", "", "",
         "False", "*True",
         "", "", "", "", "", "", "", "",
         "Correct! Contrary to popular belief, the Great Wall is not visible from space with the naked eye.",
         "Actually, this is a common myth — the Great Wall cannot be seen from space with the naked eye.",
         "10"],
        # MC with image
        ["MC",
         "Which country is home to the Eiffel Tower?",
         "media/sample_landmark.jpg", "", "",
         "*France", "Germany", "Italy", "Spain",
         "", "", "", "", "", "",
         "Well done! The Eiffel Tower is located in Paris, France.",
         "Incorrect. The Eiffel Tower is in Paris, France.",
         "10"],
        # MR
        ["MR",
         "Which of the following are programming languages? (Select all that apply)",
         "", "", "",
         "*Python", "*JavaScript", "*Rust", "HTML", "*Go",
         "", "", "", "", "",
         "Excellent! You identified all the programming languages correctly.",
         "Not quite — HTML is a markup language, not a programming language.",
         "15"],
        # TI
        ["TI",
         "What is the chemical symbol for water?\n(Type your answer in the box)",
         "", "", "",
         "H2O", "h2o",
         "", "", "", "", "", "", "", "",
         "Correct! H₂O is the chemical formula for water.",
         "Incorrect. The chemical formula for water is H₂O.",
         "10"],
        # MG
        ["MG",
         "Match each country with its capital city:",
         "media/sample_map.jpg", "", "",
         "France|Paris",
         "Japan|Tokyo",
         "Brazil|Brasília",
         "Australia|Canberra",
         "", "", "", "", "", "",
         "Perfect match! You correctly paired all countries with their capitals.",
         "Some pairs are incorrect. Review the capitals and try again.",
         "20"],
        # SEQ
        ["SEQ",
         "Place the following planets in order from closest to farthest from the Sun:",
         "", "", "",
         "Mercury", "Venus", "Earth", "Mars", "Jupiter",
         "", "", "", "", "",
         "Correct order! Mercury, Venus, Earth, Mars, Jupiter.",
         "Incorrect order. Remember: My Very Educated Mother Just Served Us Nachos.",
         "15"],
        # NUMG
        ["NUMG",
         "How many sides does a hexagon have?\n(Enter a number)",
         "", "", "",
         "=6",
         "", "", "", "", "", "", "", "", "",
         "Correct! A hexagon has exactly 6 sides.",
         "Incorrect. A hexagon has 6 sides.",
         "10"],
        # IS
        ["IS",
         "Section 2: World Geography",
         "media/sample_globe.jpg", "", "",
         "In this section you will answer questions about world geography, including capitals, landmarks, and physical features. Take your time and read each question carefully.",
         "", "", "", "", "", "", "", "", "",
         "", "", ""],
    ]

    fills = [
        make_fill(WHITE), make_fill(GRAY_BG),
        make_fill(WHITE), make_fill(GRAY_BG),
        make_fill(WHITE), make_fill(GRAY_BG),
        make_fill(WHITE), make_fill(GRAY_BG),
    ]
    type_colors = {
        "TF": "FFE8F5E9", "MC": "FFE3F2FD", "MR": "FFFCE4EC",
        "TI": "FFFFF8E1", "MG": "FFF3E5F5", "SEQ": "FFE0F2F1",
        "NUMG": "FFFBE9E7", "IS": "FFECEFF1",
    }

    for r_idx, row_data in enumerate(rows):
        for c_idx, val in enumerate(row_data):
            c = ws.cell(row=r_idx + 1, column=c_idx + 1, value=val)
            q_type = row_data[0]
            if c_idx == 0:
                c.fill = make_fill(type_colors.get(q_type, GRAY_BG))
                c.font = Font(name="Calibri", size=10, bold=True, color=DARK)
            else:
                c.fill = fills[r_idx]
                c.font = body_font()
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
            c.border = make_border()
        ws.row_dimensions[r_idx + 1].height = 40

    # Column widths
    widths = [8, 45, 25, 20, 20, 25, 25, 25, 25, 20, 20, 20, 20, 20, 20, 35, 35, 8]
    for i, w in enumerate(widths):
        set_col_width(ws, get_column_letter(i + 1), w)

    return ws


# ─── Sheet 3: Template (reference) ───────────────────────────────────────────
def build_template_sheet(wb):
    ws = wb.create_sheet("Template")
    ws.sheet_view.showGridLines = False

    headers = [
        "Question Type", "Question Text", "Image", "Video", "Audio",
        "Answer 1", "Answer 2", "Answer 3", "Answer 4", "Answer 5",
        "Answer 6", "Answer 7", "Answer 8", "Answer 9", "Answer 10",
        "Correct Feedback", "Incorrect Feedback", "Points"
    ]

    # Header row
    for c_idx, h in enumerate(headers):
        c = ws.cell(row=1, column=c_idx + 1, value=h)
        c.fill = make_fill(TEAL)
        c.font = header_font(size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = make_border(TEAL)
    ws.row_dimensions[1].height = 30

    # Reference rows
    ref_rows = [
        ["TF",   "True/False question",          "[path]","[path]","[path]", "*True","False","","","","","","","","",  "[Correct feedback]","[Incorrect feedback]","[Points]"],
        ["MC",   "Multiple Choice question",      "[path]","[path]","[path]", "*Alternative 1","Alternative 2","[Alt N]","","","","","","","",  "[Correct feedback]","[Incorrect feedback]","[Points]"],
        ["MR",   "Multiple Response question",    "[path]","[path]","[path]", "*Alternative 1","[*]Alternative 2","[Alt N]","","","","","","","","[Correct feedback]","[Incorrect feedback]","[Points]"],
        ["TI",   "Short Answer question",         "[path]","[path]","[path]", "Answer 1","[Answer N]","","","","","","","","",  "[Correct feedback]","[Incorrect feedback]","[Points]"],
        ["MG",   "Matching question",             "[path]","[path]","[path]", "Premise 1|Response 1","Premise 2|Response 2","[Premise N|Response N]","","","","","","","","[Correct feedback]","[Incorrect feedback]","[Points]"],
        ["SEQ",  "Sequence question",             "[path]","[path]","[path]", "Item 1","Item 2","Item 3","[Item N]","","","","","","",  "[Correct feedback]","[Incorrect feedback]","[Points]"],
        ["NUMG", "Numeric question",              "[path]","[path]","[path]", "[=X]","[X..Y]","[>X]","[<X]","[>=X]","[<=X]","[!=X]","","","","[Correct feedback]","[Incorrect feedback]","[Points]"],
        ["IS",   "Info Slide (not graded)",       "[path]","[path]","[path]", "Description text","","","","","","","","","",  "","",""],
    ]

    type_colors = {
        "TF": "FFE8F5E9", "MC": "FFE3F2FD", "MR": "FFFCE4EC",
        "TI": "FFFFF8E1", "MG": "FFF3E5F5", "SEQ": "FFE0F2F1",
        "NUMG": "FFFBE9E7", "IS": "FFECEFF1",
    }

    for r_idx, row_data in enumerate(ref_rows):
        for c_idx, val in enumerate(row_data):
            c = ws.cell(row=r_idx + 2, column=c_idx + 1, value=val)
            q_type = row_data[0]
            if c_idx == 0:
                c.fill = make_fill(type_colors.get(q_type, GRAY_BG))
                c.font = Font(name="Calibri", size=10, bold=True, color=DARK)
            else:
                c.fill = make_fill(GRAY_BG if r_idx % 2 == 0 else WHITE)
                c.font = Font(name="Calibri", size=9, italic=(val.startswith("[")), color="FF666666" if val.startswith("[") else DARK)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
            c.border = make_border()
        ws.row_dimensions[r_idx + 2].height = 22

    # Column widths
    widths = [8, 35, 18, 18, 18, 22, 22, 22, 18, 18, 18, 18, 18, 18, 18, 28, 28, 8]
    for i, w in enumerate(widths):
        set_col_width(ws, get_column_letter(i + 1), w)

    return ws


# ─── Create sample placeholder images ────────────────────────────────────────
def create_placeholder_image(label, width=400, height=250, color=(24, 154, 161)):
    """Create a simple colored placeholder PNG using PIL"""
    try:
        from PIL import Image, ImageDraw, ImageFont
        img = Image.new("RGB", (width, height), color=color)
        draw = ImageDraw.Draw(img)
        # Draw border
        draw.rectangle([2, 2, width-3, height-3], outline=(255,255,255), width=3)
        # Draw text
        text = f"[{label}]"
        try:
            font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 28)
        except:
            font = ImageFont.load_default()
        bbox = draw.textbbox((0, 0), text, font=font)
        tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
        draw.text(((width - tw) // 2, (height - th) // 2), text, fill=(255, 255, 255), font=font)
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        return buf.getvalue()
    except ImportError:
        # Fallback: minimal valid PNG (1x1 teal pixel)
        import struct, zlib
        def chunk(name, data):
            c = name + data
            return struct.pack(">I", len(data)) + c + struct.pack(">I", zlib.crc32(c) & 0xffffffff)
        sig = b'\x89PNG\r\n\x1a\n'
        ihdr = chunk(b'IHDR', struct.pack(">IIBBBBB", 1, 1, 8, 2, 0, 0, 0))
        raw = b'\x00' + bytes(color)
        idat = chunk(b'IDAT', zlib.compress(raw))
        iend = chunk(b'IEND', b'')
        return sig + ihdr + idat + iend


# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    output_dir = "/home/ubuntu/scorm-host/scripts"
    os.makedirs(output_dir, exist_ok=True)

    # Build workbook
    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    build_instructions_sheet(wb)
    build_questions_sheet(wb)
    build_template_sheet(wb)

    # Save workbook to bytes
    xlsx_buf = io.BytesIO()
    wb.save(xlsx_buf)
    xlsx_bytes = xlsx_buf.getvalue()

    # Sample media images
    media_files = {
        "sample_landmark.jpg": create_placeholder_image("Landmark Image", color=(24, 154, 161)),
        "sample_map.jpg":      create_placeholder_image("Map Image",      color=(74, 217, 224)),
        "sample_globe.jpg":    create_placeholder_image("Globe Image",    color=(26, 35, 50)),
    }

    # Build ZIP
    zip_path = os.path.join(output_dir, "TeachificQuizImportTemplate.zip")
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("QuizImportTemplate.xlsx", xlsx_bytes)
        for fname, data in media_files.items():
            zf.writestr(f"media/{fname}", data)

    print(f"✓ Created: {zip_path}")
    print(f"  Size: {os.path.getsize(zip_path):,} bytes")

    # Also save the xlsx standalone
    xlsx_path = os.path.join(output_dir, "QuizImportTemplate.xlsx")
    with open(xlsx_path, "wb") as f:
        f.write(xlsx_bytes)
    print(f"✓ Created: {xlsx_path}")


if __name__ == "__main__":
    main()
